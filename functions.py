import operator
import numpy as np
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import simpy

import matplotlib as mpl
import matplotlib.pyplot as plt
import openpyxl as op
import networkx as nx
import random

from deap import algorithms
from deap import base
from deap import creator
from deap import tools
from deap import gp

from statistics import mean
from PIL import Image

from operator import add
from operator import sub
from operator import mul

from simpy import Environment
from simpy import PriorityResource

from numpy.random import uniform as np_uniform
from numpy.random import choice as np_choice




def geneticprogamming(performance_measure, ref_point, opt_type, project_name, reference_point_input, nb_generations, population_size, crossover_probability, mutation_probability, max_depth_crossover, max_depth_mutation, nb_simulations):
    performance_name = ["makespan", "number tardy jobs", "total tardiness"]
    def div(left, right):
        try:
            return left / right
        except ZeroDivisionError:
            return 1

    pset = gp.PrimitiveSet("MAIN", 4)
    pset.addPrimitive(operator.add, 2)
    pset.addPrimitive(operator.sub, 2)
    pset.addPrimitive(operator.mul, 2)
    pset.addPrimitive(min, 2)
    pset.addPrimitive(max, 2)
    pset.addPrimitive(div, 2)
    #pset.addPrimitive(operator.neg, 1)
    #pset.addPrimitive(math.cos, 1)
    #pset.addPrimitive(math.sin, 1)
    #randomname = str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000))
    # Problem bei Konstante da jedes mal neuer Name generiert werden muss
    #try:
    #    pset.addEphemeralConstant(randomname, lambda: rd.uniform(-1, 1))
    #except:
    #    randomname = str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000)) + str(
    #        rd.randint(1, 100000000000000))
    #    pset.addEphemeralConstant(randomname, lambda: rd.uniform(-1, 1))
    #maybe change here to random between 0 and 10 as mentioned in overleaf
    pset.renameArguments(ARG0='PT')
    pset.renameArguments(ARG1='SPT')
    pset.renameArguments(ARG2='RPT')
    pset.renameArguments(ARG3='DD')

    if opt_type=="best":
        creator.create("FitnessMin", base.Fitness, weights=(-1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMin)
    else:
        creator.create("FitnessMax", base.Fitness, weights=(1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMax)



    toolbox = base.Toolbox()
    toolbox.register("expr", gp.genHalfAndHalf, pset=pset, min_=2, max_=6)
    toolbox.register("individual", tools.initIterate, creator.Individual, toolbox.expr)
    toolbox.register("population", tools.initRepeat, list, toolbox.individual)
    toolbox.register("compile", gp.compile, pset=pset)


    def evalSymbReg(individual):
        func = toolbox.compile(expr=individual)

        def visualize(results):
            schedule = pd.DataFrame(results)
            JOBS = sorted(list(schedule['Job'].unique()))
            MACHINES = sorted(list(schedule['Machine'].unique()))
            makespan = schedule['Finish'].max()

            bar_style = {'alpha': 1.0, 'lw': 25, 'solid_capstyle': 'butt'}
            text_style = {'color': 'white', 'weight': 'bold', 'ha': 'center', 'va': 'center'}
            colors = mpl.cm.Dark2.colors

            schedule.sort_values(by=['Job', 'Start'])
            schedule.set_index(['Job', 'Machine'], inplace=True)

            fig, ax = plt.subplots(2, 1, figsize=(12, 5 + (len(JOBS) + len(MACHINES)) / 4))

            for jdx, j in enumerate(JOBS, 1):
                for mdx, m in enumerate(MACHINES, 1):
                    if (j, m) in schedule.index:
                        xs = schedule.loc[(j, m), 'Start']
                        xf = schedule.loc[(j, m), 'Finish']
                        ax[0].plot([xs, xf], [jdx] * 2, c=colors[mdx % 7], **bar_style)
                        ax[0].text((xs + xf) / 2, jdx, m, **text_style)
                        ax[1].plot([xs, xf], [mdx] * 2, c=colors[jdx % 7], **bar_style)
                        ax[1].text((xs + xf) / 2, mdx, j, **text_style)

            ax[0].set_title('Job Schedule')
            ax[0].set_ylabel('Job')
            ax[1].set_title('Machine Schedule')
            ax[1].set_ylabel('Machine')

            for idx, s in enumerate([JOBS, MACHINES]):
                ax[idx].set_ylim(0.5, len(s) + 0.5)
                ax[idx].set_yticks(range(1, 1 + len(s)))
                ax[idx].set_yticklabels(s)
                ax[idx].text(makespan, ax[idx].get_ylim()[0] - 0.2, "{0:0.1f}".format(makespan), ha='center', va='top')
                ax[idx].plot([makespan] * 2, ax[idx].get_ylim(), 'r--')
                ax[idx].set_xlabel('Time')
                ax[idx].grid(True)

            fig.tight_layout()
            plt.show()

        def source(env, number, machine):
            """Source generates jobs randomly"""
            for i in range(number):
                processing_time = np.random.uniform(avg_processing_time - delta_processing_time,
                                                    avg_processing_time + delta_processing_time)
                if i == 0:
                    release_time = np.random.uniform(avg_job_interarrival_time - delta_release_date,
                                                     avg_job_interarrival_time + delta_release_date)
                else:
                    release_time = release_time_list[i - 1] + np.random.uniform(
                        avg_job_interarrival_time - delta_release_date,
                        avg_job_interarrival_time + delta_release_date)

                due_date = release_time + k * processing_time
                due_date_list.append(due_date)
                processing_time_list.append(processing_time)
                release_time_list.append(release_time)
                c = job(env, f'Job {i + 1}', machine, processing_time=processing_time,
                        total_processing_time=total_processing_time)
                env.process(c)
                t = 0
                yield env.timeout(t)

        def job(env, name, machine, processing_time, total_processing_time, remaining_processing_time, due_date,
                release_date):
            """Job arrives, is served and leaves."""
            # arrive = release_time
            # print('%7.4f %s: Arrived' % (arrive, name))

            with machine.request(priority=func(processing_time, total_processing_time, remaining_processing_time, due_date)) as req:
                yield req
                # wait = env.now - arrive # waiting time of job
                job_start = env.now

                # We got to the counter
                # print('%7.4f %s: Waited %6.3f' % (env.now, name, wait))

                yield env.timeout(processing_time)
                # print('%7.4f %s: Finished' % (env.now, name))

                # Flow Time
                # flow_time.append(env.now-arrive)
                remaining_processing_time -= processing_time
                job_prec = name
                machine_prec = machine_list.index(machine)
                schedule.append([job_prec, machine_prec])
                #results.append({'Job': f'J{job_prec}',
                #                'Machine': f'M{machine_prec}',
                #                'Start': job_start,
                #                'Duration': processing_time,
                #                'Finish': env.now})
                job_finished = 'Yes'
                for j in range(10):
                    for m in range(10):
                        if TASKS[j, m]['prec'] == (job_prec, machine_prec):
                            machine = machine_list[m]
                            processing_time = TASKS[j, m]['dur']
                            job_finished = 'No'
                            env.process(
                                job(env, j, machine, processing_time, total_processing_time, remaining_processing_time,
                                    due_date, release_date))
                if job_finished == 'Yes':
                    # Completion time
                    completion_time.append(env.now)
                    # Tardiness of job
                    tardiness.append(max(env.now - due_date, 0))
                    # Tardy jobs
                    if max(env.now - due_date, 0) > 0:
                        tardy_jobs.append(1)
                    else:
                        tardy_jobs.append(0)

        number_simulations = nb_simulations
        avg_makespan_list = []
        avg_mean_tardiness_list = []
        avg_max_tardiness_list = []
        avg_total_tardiness_list = []
        avg_number_tardy_jobs_list = []

        for simulations in range(number_simulations):
            np.random.seed(10)
            # random job generator
            number_machines = 10
            n_jobs = 10
            duedate_tightness = 1.5
            TASKS = {}
            release_date_list = []
            due_date_list_jobs = []
            total_processing_time = []
            for i in range(n_jobs):
                prec = None
                release_time = np.random.uniform(0, 40)
                sum_proc_time = 0
                allowed_values = list(range(0, 10))
                for m in range(number_machines):
                    dur = np.random.uniform(number_machines / 2, number_machines * 2)
                    sum_proc_time += dur
                    machine = np.random.choice(allowed_values)
                    task = (i, machine)
                    TASKS[task] = {'dur': int(dur), 'prec': prec}
                    prec = task
                    allowed_values.remove(machine)
                due_date = release_time + duedate_tightness * sum_proc_time
                release_date_list.append(release_time)
                due_date_list_jobs.append(due_date)

            for j in range(n_jobs):
                total_processing_time_current_job = 0
                for m in range(number_machines):
                    total_processing_time_current_job += TASKS[j, m]['dur']
                total_processing_time.append(total_processing_time_current_job)

            # print(TASKS['J0', 'M0']['prec'])
            # print(pd.DataFrame(TASKS).T)

            number_simulations = 1
            R_processing_time = 0.4
            avg_processing_time = 10
            processing_time = []
            duedate_tightness = 2
            duedate_variability = 0.3
            machine_utilization = 0.7
            job_interarrival_tightness = 1
            schedule = []
            release_time = []
            tardiness = []
            tardy_jobs = []
            completion_time = []
            flow_time = []
            due_date_list = []
            release_time_list = []
            processing_time_list = []
            results = []
            remaining_processing_time_current_job = []

            n_jobs_original = n_jobs
            delta_processing_time = R_processing_time * avg_processing_time
            avg_job_interarrival_time = avg_processing_time / machine_utilization
            delta_release_date = job_interarrival_tightness * avg_job_interarrival_time
            release_time.append(np.random.uniform(avg_job_interarrival_time - delta_release_date,
                                                  avg_job_interarrival_time + delta_release_date))
            delta_duedate = duedate_tightness * duedate_variability
            k = np.random.uniform(duedate_tightness - delta_duedate, duedate_tightness + delta_duedate)

            env = simpy.Environment()

            # Start processes and run
            machine_0 = simpy.PriorityResource(env, capacity=1)
            machine_1 = simpy.PriorityResource(env, capacity=1)
            machine_2 = simpy.PriorityResource(env, capacity=1)
            machine_3 = simpy.PriorityResource(env, capacity=1)
            machine_4 = simpy.PriorityResource(env, capacity=1)
            machine_5 = simpy.PriorityResource(env, capacity=1)
            machine_6 = simpy.PriorityResource(env, capacity=1)
            machine_7 = simpy.PriorityResource(env, capacity=1)
            machine_8 = simpy.PriorityResource(env, capacity=1)
            machine_9 = simpy.PriorityResource(env, capacity=1)
            # machine = simpy.Resource(env, capacity=1)

            machine_list = [machine_0, machine_1, machine_2, machine_3, machine_4, machine_5, machine_6, machine_7,
                            machine_8, machine_9]

            for j in range(n_jobs):
                for m in range(number_machines):
                    if TASKS[j, m]['prec'] == None:
                        current_machine = machine_list[m]
                        processing_time_new = float(TASKS[j, m]['dur'])
                        job_new = j
                        total_processing_time_current_job = total_processing_time[j]
                        remaining_processing_time_current_job = total_processing_time_current_job
                        release_date_current_job = release_date_list[j]
                        due_date_current_job = due_date_list_jobs[j]
                        env.process(job(env, job_new, current_machine, processing_time=processing_time_new,
                                        total_processing_time=total_processing_time_current_job,
                                        remaining_processing_time=remaining_processing_time_current_job,
                                        due_date=due_date_current_job, release_date=release_date_current_job))

            env.run()

            # for i in range(number_simulations):
            #    env.process(source(env, n_jobs, machine_1))
            #    env.run()

            # Post processing
            # calculate and performance measures of the current simulation run
            total_tardiness = sum(tardiness)
            mean_tardiness = mean(tardiness)
            max_tardiness = max(tardiness)
            number_tardy_jobs = sum(tardy_jobs)
            makespan = max(completion_time)
            # mean_flow_time = mean(flow_time)
            # max_flow_time = max(flow_time)
            # print('Release Time')
            # print(release_time_list)
            # print('processing time')
            # print(processing_time_list)
            # print('Due Dates')
            # print(due_date_list)
            # print(f'Total Tardiness: {total_tardiness}')
            # print(f'Mean Tardiness: {mean_tardiness}')
            # print(f'Max Tardiness: {max_tardiness}')
            # print(f'Number of tardy Jobs: {number_tardy_jobs}')
            # print(completion_time)
            # print(f'Makespan: {makespan}')
            # print(f'Mean flow time: {mean_flow_time}')
            # print(f'Max flow time: {max_flow_time}')
            # print(results)

            # add performance measures of current simulation run to the list for all runs
            avg_makespan_list.append(makespan)
            avg_mean_tardiness_list.append(mean_tardiness)
            avg_max_tardiness_list.append(max_tardiness)
            avg_total_tardiness_list.append(total_tardiness)
            avg_number_tardy_jobs_list.append(number_tardy_jobs)

            # visualize(results)

        # calculate and print the performance measures after all simulation runs
        avg_makespan = mean(avg_makespan_list)
        avg_mean_tardiness = mean(avg_mean_tardiness_list)
        avg_max_tardiness = mean(avg_max_tardiness_list)
        avg_total_tardiness = mean(avg_total_tardiness_list)
        avg_number_tardy_jobs = mean(avg_number_tardy_jobs_list)
        performance = [avg_makespan, avg_number_tardy_jobs, avg_total_tardiness]
        lambda_performance = [0.33, 0.33, 0.33]
        reference_point = reference_point_input
        roh = 0.0001
        # calculate the fitness value according to the achievement scalarizing function
        if ref_point==None:
            fitness = performance[performance_measure]
        else:
            fitness = max(lambda_performance[i]*(performance[i]-reference_point[i]) for i in range(0,len(reference_point))) + roh * sum(lambda_performance[i]*(performance[i]-reference_point[i]) for i in range(0,len(reference_point)))
        return fitness,


    toolbox.register("evaluate", evalSymbReg)
    toolbox.register("select", tools.selTournament, tournsize=5)
    toolbox.register("mate", gp.cxOnePoint)
    toolbox.register("expr_mut", gp.genFull, min_=0, max_=2)
    toolbox.register("mutate", gp.mutUniform, expr=toolbox.expr_mut, pset=pset)

    toolbox.decorate("mate", gp.staticLimit(key=operator.attrgetter("height"), max_value=max_depth_crossover))
    toolbox.decorate("mutate", gp.staticLimit(key=operator.attrgetter("height"), max_value=max_depth_mutation))

    random.seed(318)

    pop = toolbox.population(n=population_size)
    hof = tools.HallOfFame(1)

    stats_fit = tools.Statistics(lambda ind: ind.fitness.values)
    stats_size = tools.Statistics(len)
    mstats = tools.MultiStatistics(fitness=stats_fit, size=stats_size)
    mstats.register("avg", np.mean)
    mstats.register("std", np.std)
    mstats.register("min", np.min)
    mstats.register("max", np.max)

    pop, log = algorithms.eaSimple(pop, toolbox, crossover_probability, mutation_probability, nb_generations, stats=mstats,
                                   halloffame=hof, verbose=True)

    # extract statistics:
    avgFitnessValues  = log.chapters['fitness'].select("avg")
    minFitnessValues = log.chapters['fitness'].select("min")
    maxFitnessValues = log.chapters['fitness'].select("max")
    stdFitnessValues = log.chapters['fitness'].select("std")
    nb_generation = log.select("gen")
    nevals = log.select('nevals')

    # plot statistics:
    # sns.set_style("whitegrid")
    fig, ax = plt.subplots()
    #ax.boxplot(minFitnessValues)
    mins = np.array(minFitnessValues)
    maxes = np.array(maxFitnessValues)
    means = np.array(avgFitnessValues)
    std = np.array(stdFitnessValues)
    gen = np.array(nb_generation)

    # create stacked errorbars:

    #ax.scatter(nb_generation, avgFitnessValues)
    ax.errorbar(gen, means, std, fmt='ok', lw=3)
    ax.errorbar(gen, means, [means - mins, maxes - means],
                 fmt='.k', ecolor='gray', lw=1)
    ax.set_xlabel('Generation')
    ax.set_ylabel('Average Fitness')
    st.header('Evolution process')
    st.pyplot(fig)

    #print("best solution:")
    #print(hof[0])
    best = hof.items[0]
    best_fitness = best.fitness.values[0]
    st.header('Dispatching rule of the best solution')
    st.write(str(best))

    # create visualisation of tree of the final solution
    nodes, edges, labels = gp.graph(hof.items[0])
    fig, ax = plt.subplots()
    g = nx.Graph()
    g.add_nodes_from(nodes)
    g.add_edges_from(edges)
    pos = nx.nx_agraph.graphviz_layout(g, prog='dot')
    nx.draw_networkx_nodes(g, pos)
    nx.draw_networkx_edges(g, pos)
    nx.draw_networkx_labels(g, pos, labels, font_size=6, font_color="whitesmoke")
    path = 'C:/Users/yanni/PycharmProjects/geneticalgorithm/html_files'
    plt.savefig(f'{path}/nx_graph.png')
    image = Image.open(f'{path}/nx_graph.png')
    st.header('Primitive tree of the best solution')
    st.image(image,  use_column_width=True)

    wb = op.load_workbook(project_name + '.xlsx')
    ws_ideal_nadir =  wb["ideal_nadir"]
    ws_ref_point = wb["ref_point"]
    max_row = ws_ref_point.max_row
    if ref_point==None:
        st.header(performance_name[performance_measure])
        st.write(opt_type + " solution")
        st.write(best_fitness)
        if opt_type == 'best':
            ws_ideal_nadir['B'+str(performance_measure+2)] = best_fitness
        else:
            ws_ideal_nadir['C' + str(performance_measure + 2)] = best_fitness
    else:
        st.header('Best solution')
        st.write(best_fitness)
        ws_ref_point['A' + str(max_row + 1)] = str(reference_point_input)
        #ws_ref_point['A'+str(max_row+1)] = "["+str(reference_point_input[0]) +", " + str(reference_point_input[1]) +", "+ str(reference_point_input[2]) +"]"
        ws_ref_point['B' + str(max_row + 1)] = best_fitness
        ws_ref_point['C' + str(max_row + 1)] = str(best)

    print('best fitness slow \t\t', best_fitness)

    wb.save(project_name+'.xlsx')
    wb.close()
    #st.write(pop)
    #st.write(log)
    #st.stop()


def geneticprogamming_speedup(performance_measure, ref_point, opt_type, project_name, reference_point_input, nb_generations, population_size, crossover_probability, mutation_probability, max_depth_crossover, max_depth_mutation, nb_simulations):
    performance_name = ["makespan", "number tardy jobs", "total tardiness"]
    def div(left, right):
        try:
            return left / right
        except ZeroDivisionError:
            return 1

    pset = gp.PrimitiveSet("MAIN", 4)
    pset.addPrimitive(add, 2)
    pset.addPrimitive(sub, 2)
    pset.addPrimitive(mul, 2)
    pset.addPrimitive(min, 2)
    pset.addPrimitive(max, 2)
    pset.addPrimitive(div, 2)

    #randomname = str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000))
    # Problem bei Konstante da jedes mal neuer Name generiert werden muss
    #try:
    #    pset.addEphemeralConstant(randomname, lambda: rd.uniform(-1, 1))
    #except:
    #    randomname = str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000)) + str(
    #        rd.randint(1, 100000000000000))
    #    pset.addEphemeralConstant(randomname, lambda: rd.uniform(-1, 1))
    #maybe change here to random between 0 and 10 as mentioned in overleaf
    pset.renameArguments(ARG0='PT')
    pset.renameArguments(ARG1='SPT')
    pset.renameArguments(ARG2='RPT')
    pset.renameArguments(ARG3='DD')

    if opt_type=="best":
        creator.create("FitnessMin", base.Fitness, weights=(-1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMin)
    else:
        creator.create("FitnessMax", base.Fitness, weights=(1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMax)



    toolbox = base.Toolbox()
    toolbox.register("expr", gp.genHalfAndHalf, pset=pset, min_=2, max_=6)
    toolbox.register("individual", tools.initIterate, creator.Individual, toolbox.expr)
    toolbox.register("population", tools.initRepeat, list, toolbox.individual)
    toolbox.register("compile", gp.compile, pset=pset)


    def evalSymbReg(individual):
        func = toolbox.compile(expr=individual)

        def job(env, name, machine, processing_time, total_processing_time, remaining_processing_time, due_date,
                release_date):
            """Job arrives, is served and leaves."""

            with machine.request(priority=func(processing_time, total_processing_time, remaining_processing_time, due_date)) as req:
                yield req
                job_start = env.now
                # We got to the counter

                yield env.timeout(processing_time)
                remaining_processing_time -= processing_time
                job_prec = name
                machine_prec = machine_list.index(machine)
                schedule.append([job_prec, machine_prec])
                job_finished = 'Yes'
                for j in range(10):
                    for m in range(10):
                        if TASKS[j, m]['prec'] == (job_prec, machine_prec):
                            machine = machine_list[m]
                            processing_time = TASKS[j, m]['dur']
                            job_finished = 'No'
                            env.process(
                                job(env, j, machine, processing_time, total_processing_time, remaining_processing_time,
                                    due_date, release_date))
                if job_finished == 'Yes':
                    # Completion time
                    completion_time.append(env.now)
                    # Tardiness of job
                    tardiness.append(max(env.now - due_date, 0))
                    # Tardy jobs
                    if max(env.now - due_date, 0) > 0:
                        tardy_jobs.append(1)
                    else:
                        tardy_jobs.append(0)

        number_machines = 10
        n_jobs = 10
        number_simulations = nb_simulations

        #define all lists
        avg_makespan_list = [i for i in range(nb_simulations)]
        #avg_makespan_list = np.array(i for i in range(nb_simulations))
        #print(avg_makespan_list)
        avg_mean_tardiness_list = [i for i in range(nb_simulations)]
        #print(avg_mean_tardiness_list)
        avg_max_tardiness_list = [i for i in range(nb_simulations)]
        avg_total_tardiness_list = [i for i in range(nb_simulations)]
        avg_number_tardy_jobs_list = [i for i in range(nb_simulations)]
        release_date_list = [i for i in range(n_jobs)]
        due_date_list_jobs = [i for i in range(n_jobs)]
        total_processing_time = [i for i in range(n_jobs)]

        for simulations in range(number_simulations):
            np.random.seed(10)
            # random job generator
            duedate_tightness = 1.5
            TASKS = {}
            for i in range(n_jobs):
                prec = None
                release_time = np_uniform(0, 40)
                sum_proc_time = 0
                allowed_values = list(range(0, number_machines))
                for m in range(number_machines):
                    dur = np_uniform(number_machines / 2, number_machines * 2)
                    sum_proc_time += dur
                    machine = np_choice(allowed_values)
                    task = (i, machine)
                    TASKS[task] = {'dur': int(dur), 'prec': prec}
                    prec = task
                    allowed_values.remove(machine)
                due_date = release_time + duedate_tightness * sum_proc_time
                release_date_list[i] = release_time
                due_date_list_jobs[i] = due_date

            for j in range(n_jobs):
                total_processing_time_current_job = 0
                for m in range(number_machines):
                    total_processing_time_current_job += TASKS[j, m]['dur']
                total_processing_time[j] = total_processing_time_current_job

            schedule = []
            tardiness = []
            tardy_jobs = []
            completion_time = []

            env = Environment()

            # Start processes and run
            machine_0 = PriorityResource(env, capacity=1)
            machine_1 = PriorityResource(env, capacity=1)
            machine_2 = PriorityResource(env, capacity=1)
            machine_3 = PriorityResource(env, capacity=1)
            machine_4 = PriorityResource(env, capacity=1)
            machine_5 = PriorityResource(env, capacity=1)
            machine_6 = PriorityResource(env, capacity=1)
            machine_7 = PriorityResource(env, capacity=1)
            machine_8 = PriorityResource(env, capacity=1)
            machine_9 = PriorityResource(env, capacity=1)
            # machine = simpy.Resource(env, capacity=1)

            machine_list = [machine_0, machine_1, machine_2, machine_3, machine_4, machine_5, machine_6, machine_7,
                            machine_8, machine_9]

            for j in range(n_jobs):
                for m in range(number_machines):
                    if TASKS[j, m]['prec'] == None:
                        current_machine = machine_list[m]
                        processing_time_new = float(TASKS[j, m]['dur'])
                        job_new = j
                        total_processing_time_current_job = total_processing_time[j]
                        remaining_processing_time_current_job = total_processing_time_current_job
                        release_date_current_job = release_date_list[j]
                        due_date_current_job = due_date_list_jobs[j]
                        env.process(job(env, job_new, current_machine, processing_time=processing_time_new,
                                        total_processing_time=total_processing_time_current_job,
                                        remaining_processing_time=remaining_processing_time_current_job,
                                        due_date=due_date_current_job, release_date=release_date_current_job))

            env.run()

            # Post processing
            # calculate and performance measures of the current simulation run
            total_tardiness = np.sum(tardiness)
            mean_tardiness = np.mean(tardiness)
            max_tardiness = np.amax(tardiness)
            number_tardy_jobs = np.sum(tardy_jobs)
            makespan = np.amax(completion_time)

            # add performance measures of current simulation run to the list for all runs
            avg_makespan_list[simulations] = makespan
            avg_mean_tardiness_list[simulations] = mean_tardiness
            avg_max_tardiness_list[simulations] = max_tardiness
            avg_total_tardiness_list[simulations] = total_tardiness
            avg_number_tardy_jobs_list[simulations] = number_tardy_jobs

            # visualize(results)

        # calculate and print the performance measures after all simulation runs
        avg_makespan = np.mean(avg_makespan_list)
        avg_total_tardiness = np.mean(avg_total_tardiness_list)
        avg_number_tardy_jobs = np.mean(avg_number_tardy_jobs_list)
        performance = [avg_makespan, avg_number_tardy_jobs, avg_total_tardiness]
        lambda_performance = [0.33, 0.33, 0.33]
        reference_point = reference_point_input
        roh = 0.0001
        # calculate the fitness value according to the achievement scalarizing function
        if ref_point==None:
            fitness = performance[performance_measure]
        else:
            fitness = max(lambda_performance[i]*(performance[i]-reference_point[i]) for i in range(0,len(reference_point))) + roh * sum(lambda_performance[i]*(performance[i]-reference_point[i]) for i in range(0,len(reference_point)))
        return fitness,


    toolbox.register("evaluate", evalSymbReg)
    toolbox.register("select", tools.selTournament, tournsize=5)
    toolbox.register("mate", gp.cxOnePoint)
    toolbox.register("expr_mut", gp.genFull, min_=0, max_=2)
    toolbox.register("mutate", gp.mutUniform, expr=toolbox.expr_mut, pset=pset)

    toolbox.decorate("mate", gp.staticLimit(key=operator.attrgetter("height"), max_value=max_depth_crossover))
    toolbox.decorate("mutate", gp.staticLimit(key=operator.attrgetter("height"), max_value=max_depth_mutation))

    random.seed(318)

    pop = toolbox.population(n=population_size)
    hof = tools.HallOfFame(1)

    stats_fit = tools.Statistics(lambda ind: ind.fitness.values)
    stats_size = tools.Statistics(len)
    mstats = tools.MultiStatistics(fitness=stats_fit, size=stats_size)
    mstats.register("avg", np.mean)
    mstats.register("std", np.std)
    mstats.register("min", np.min)
    mstats.register("max", np.max)

    pop, log = algorithms.eaSimple(pop, toolbox, crossover_probability, mutation_probability, nb_generations, stats=mstats,
                                   halloffame=hof, verbose=True)

    # extract statistics:
    avgFitnessValues  = log.chapters['fitness'].select("avg")
    minFitnessValues = log.chapters['fitness'].select("min")
    maxFitnessValues = log.chapters['fitness'].select("max")
    stdFitnessValues = log.chapters['fitness'].select("std")
    nb_generation = log.select("gen")
    # plot statistics:
    # sns.set_style("whitegrid")
    fig, ax = plt.subplots()
    #ax.boxplot(minFitnessValues)
    mins = np.array(minFitnessValues)
    maxes = np.array(maxFitnessValues)
    means = np.array(avgFitnessValues)
    std = np.array(stdFitnessValues)
    gen = np.array(nb_generation)

    # create stacked errorbars:

    #ax.scatter(nb_generation, avgFitnessValues)
    ax.errorbar(gen, means, std, fmt='ok', lw=3)
    ax.errorbar(gen, means, [means - mins, maxes - means],
                 fmt='.k', ecolor='gray', lw=1)
    ax.set_xlabel('Generation')
    ax.set_ylabel('Average Fitness')
    st.header('Evolution process')
    st.pyplot(fig)

    #print("best solution:")
    #print(hof[0])
    best = hof.items[0]
    best_fitness = best.fitness.values[0]
    st.header('Dispatching rule of the best solution')
    st.write(str(best))

    # create visualisation of tree of the final solution
    nodes, edges, labels = gp.graph(hof.items[0])
    fig, ax = plt.subplots()
    g = nx.Graph()
    g.add_nodes_from(nodes)
    g.add_edges_from(edges)
    pos = nx.nx_agraph.graphviz_layout(g, prog='dot')
    nx.draw_networkx_nodes(g, pos)
    nx.draw_networkx_edges(g, pos)
    nx.draw_networkx_labels(g, pos, labels, font_size=6, font_color="whitesmoke")
    path = 'C:/Users/yanni/PycharmProjects/geneticalgorithm/html_files'
    plt.savefig(f'{path}/nx_graph.png')
    image = Image.open(f'{path}/nx_graph.png')
    st.header('Primitive tree of the best solution')
    st.image(image,  use_column_width=True)

    wb = op.load_workbook(project_name + '.xlsx')
    ws_ideal_nadir =  wb["ideal_nadir"]
    ws_ref_point = wb["ref_point"]
    max_row = ws_ref_point.max_row
    if ref_point==None:
        st.header(performance_name[performance_measure])
        st.write(opt_type + " solution")
        st.write(best_fitness)
        if opt_type == 'best':
            ws_ideal_nadir['B'+str(performance_measure+2)] = best_fitness
        else:
            ws_ideal_nadir['C' + str(performance_measure + 2)] = best_fitness
    else:
        st.header('Best solution')
        st.write(best_fitness)
        ws_ref_point['A' + str(max_row + 1)] = str(reference_point_input)
        #ws_ref_point['A'+str(max_row+1)] = "["+str(reference_point_input[0]) +", " + str(reference_point_input[1]) +", "+ str(reference_point_input[2]) +"]"
        ws_ref_point['B' + str(max_row + 1)] = best_fitness
        ws_ref_point['C' + str(max_row + 1)] = str(best)

    print('best fitness fast \t\t', best_fitness)

    wb.save(project_name+'.xlsx')
    wb.close()


def start_new_project(project_name, performance_name):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "ideal_nadir"
    for i in range(len(performance_name)):
        ws['A'+str(i+2)] = performance_name[i]
    ws['B1'] = 'ideal point'
    ws['C1'] = 'nadir point'
    ws1 = wb.create_sheet("ref_point")  # insert at the end (default)
    ws1['A1'] = 'reference point'
    ws1['B1'] = 'best fitness'
    ws1['C1'] = 'best function'
    wb.save(project_name+'.xlsx')
    wb.close()

def verification(performance_measure, ref_point, reference_point_input, best_func):

    performance_name = ["makespan", "number tardy jobs", "total tardiness"]
    def div(left, right):
        try:
            return left / right
        except ZeroDivisionError:
            return 1

    pset = gp.PrimitiveSet("MAIN", 4)
    pset.addPrimitive(operator.add, 2)
    pset.addPrimitive(operator.sub, 2)
    pset.addPrimitive(operator.mul, 2)
    pset.addPrimitive(min, 2)
    pset.addPrimitive(max, 2)
    pset.addPrimitive(div, 2)
    #pset.addPrimitive(operator.neg, 1)
    #pset.addPrimitive(math.cos, 1)
    #pset.addPrimitive(math.sin, 1)
    #randomname = str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000))
    # Problem bei Konstante da jedes mal neuer Name generiert werden muss
    #try:
    #    pset.addEphemeralConstant(randomname, lambda: rd.uniform(-1, 1))
    #except:
    #    randomname = str(rd.randint(1, 100000000000000)) + str(rd.randint(1, 100000000000000)) + str(
    #        rd.randint(1, 100000000000000))
    #    pset.addEphemeralConstant(randomname, lambda: rd.uniform(-1, 1))
    #maybe change here to random between 0 and 10 as mentioned in overleaf
    pset.renameArguments(ARG0='PT')
    pset.renameArguments(ARG1='SPT')
    pset.renameArguments(ARG2='RPT')
    pset.renameArguments(ARG3='DD')
    opt_type = 'best'
    if opt_type=="best":
        creator.create("FitnessMin", base.Fitness, weights=(-1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMin)
    else:
        creator.create("FitnessMax", base.Fitness, weights=(1.0,))
        creator.create("Individual", gp.PrimitiveTree, fitness=creator.FitnessMax)



    toolbox = base.Toolbox()
    toolbox.register("expr", gp.genHalfAndHalf, pset=pset, min_=2, max_=6)
    toolbox.register("individual", tools.initIterate, creator.Individual, toolbox.expr)
    toolbox.register("population", tools.initRepeat, list, toolbox.individual)
    toolbox.register("compile", gp.compile, pset=pset)

    func = toolbox.compile(expr=best_func)

    def visualize(results):
        schedule = pd.DataFrame(results)
        JOBS = sorted(list(schedule['Job'].unique()))
        MACHINES = sorted(list(schedule['Machine'].unique()))
        makespan = schedule['Finish'].max()

        bar_style = {'alpha': 1.0, 'lw': 25, 'solid_capstyle': 'butt'}
        text_style = {'color': 'white', 'weight': 'bold', 'ha': 'center', 'va': 'center'}
        colors = mpl.cm.Dark2.colors

        schedule.sort_values(by=['Job', 'Start'])
        schedule.set_index(['Job', 'Machine'], inplace=True)

        fig, ax = plt.subplots(2, 1, figsize=(12, 5 + (len(JOBS) + len(MACHINES)) / 4))

        for jdx, j in enumerate(JOBS, 1):
            for mdx, m in enumerate(MACHINES, 1):
                if (j, m) in schedule.index:
                    xs = schedule.loc[(j, m), 'Start']
                    xf = schedule.loc[(j, m), 'Finish']
                    ax[0].plot([xs, xf], [jdx] * 2, c=colors[mdx % 7], **bar_style)
                    ax[0].text((xs + xf) / 2, jdx, m, **text_style)
                    ax[1].plot([xs, xf], [mdx] * 2, c=colors[jdx % 7], **bar_style)
                    ax[1].text((xs + xf) / 2, mdx, j, **text_style)

        ax[0].set_title('Job Schedule')
        ax[0].set_ylabel('Job')
        ax[1].set_title('Machine Schedule')
        ax[1].set_ylabel('Machine')

        for idx, s in enumerate([JOBS, MACHINES]):
            ax[idx].set_ylim(0.5, len(s) + 0.5)
            ax[idx].set_yticks(range(1, 1 + len(s)))
            ax[idx].set_yticklabels(s)
            ax[idx].text(makespan, ax[idx].get_ylim()[0] - 0.2, "{0:0.1f}".format(makespan), ha='center', va='top')
            ax[idx].plot([makespan] * 2, ax[idx].get_ylim(), 'r--')
            ax[idx].set_xlabel('Time')
            ax[idx].grid(True)

        fig.tight_layout()
        #plt.show()
        st.pyplot(fig)

    def source(env, number, machine):
        """Source generates jobs randomly"""
        for i in range(number):
            processing_time = np.random.uniform(avg_processing_time - delta_processing_time,
                                                avg_processing_time + delta_processing_time)
            if i == 0:
                release_time = np.random.uniform(avg_job_interarrival_time - delta_release_date,
                                                 avg_job_interarrival_time + delta_release_date)
            else:
                release_time = release_time_list[i - 1] + np.random.uniform(
                    avg_job_interarrival_time - delta_release_date,
                    avg_job_interarrival_time + delta_release_date)

            due_date = release_time + k * processing_time
            due_date_list.append(due_date)
            processing_time_list.append(processing_time)
            release_time_list.append(release_time)
            c = job(env, f'Job {i + 1}', machine, processing_time=processing_time,
                    total_processing_time=total_processing_time)
            env.process(c)
            t = 0
            yield env.timeout(t)

    def job(env, name, machine, processing_time, total_processing_time, remaining_processing_time, due_date,
            release_date):
        """Job arrives, is served and leaves."""
        # arrive = release_time
        # print('%7.4f %s: Arrived' % (arrive, name))

        with machine.request(
                priority=func(processing_time, total_processing_time, remaining_processing_time, due_date)) as req:
            yield req
            # wait = env.now - arrive # waiting time of job
            job_start = env.now

            # We got to the counter
            # print('%7.4f %s: Waited %6.3f' % (env.now, name, wait))

            yield env.timeout(processing_time)
            # print('%7.4f %s: Finished' % (env.now, name))

            # Flow Time
            # flow_time.append(env.now-arrive)
            remaining_processing_time -= processing_time
            job_prec = name
            machine_prec = machine_list.index(machine)
            schedule.append([job_prec, machine_prec])
            results.append({'Job': f'J{job_prec}',
                            'Machine': f'M{machine_prec}',
                            'Start': job_start,
                            'Duration': processing_time,
                            'Finish': env.now})
            job_finished = 'Yes'
            for j in range(10):
                for m in range(10):
                    if TASKS[j, m]['prec'] == (job_prec, machine_prec):
                        machine = machine_list[m]
                        processing_time = TASKS[j, m]['dur']
                        job_finished = 'No'
                        env.process(
                            job(env, j, machine, processing_time, total_processing_time, remaining_processing_time,
                                due_date, release_date))
            if job_finished == 'Yes':
                # Completion time
                completion_time.append(env.now)
                # Tardiness of job
                tardiness.append(max(env.now - due_date, 0))
                # Tardy jobs
                if max(env.now - due_date, 0) > 0:
                    tardy_jobs.append(1)
                else:
                    tardy_jobs.append(0)

    number_simulations = 50
    avg_makespan_list = []
    avg_mean_tardiness_list = []
    avg_max_tardiness_list = []
    avg_total_tardiness_list = []
    avg_number_tardy_jobs_list = []

    for simulations in range(number_simulations):
        np.random.seed(10)
        # random job generator
        number_machines = 10
        n_jobs = 10
        duedate_tightness = 1.5
        TASKS = {}
        release_date_list = []
        due_date_list_jobs = []
        total_processing_time = []
        for i in range(n_jobs):
            prec = None
            release_time = np.random.uniform(0, 40)
            sum_proc_time = 0
            allowed_values = list(range(0, number_machines))
            for m in range(number_machines):
                dur = np.random.uniform(number_machines / 2, number_machines * 2)
                sum_proc_time += dur
                machine = np.random.choice(allowed_values)
                task = (i, machine)
                TASKS[task] = {'dur': int(dur), 'prec': prec}
                prec = task
                allowed_values.remove(machine)
            due_date = release_time + duedate_tightness * sum_proc_time
            release_date_list.append(release_time)
            due_date_list_jobs.append(due_date)

        for j in range(n_jobs):
            total_processing_time_current_job = 0
            for m in range(number_machines):
                total_processing_time_current_job += TASKS[j, m]['dur']
            total_processing_time.append(total_processing_time_current_job)

        # print(TASKS['J0', 'M0']['prec'])
        # print(pd.DataFrame(TASKS).T)

        number_simulations = 1
        R_processing_time = 0.4
        avg_processing_time = 10
        processing_time = []
        duedate_tightness = 2
        duedate_variability = 0.3
        machine_utilization = 0.7
        job_interarrival_tightness = 1
        schedule = []
        release_time = []
        tardiness = []
        tardy_jobs = []
        completion_time = []
        flow_time = []
        due_date_list = []
        release_time_list = []
        processing_time_list = []
        results = []
        remaining_processing_time_current_job = []

        n_jobs_original = n_jobs
        delta_processing_time = R_processing_time * avg_processing_time
        avg_job_interarrival_time = avg_processing_time / machine_utilization
        delta_release_date = job_interarrival_tightness * avg_job_interarrival_time
        release_time.append(np.random.uniform(avg_job_interarrival_time - delta_release_date,
                                              avg_job_interarrival_time + delta_release_date))
        delta_duedate = duedate_tightness * duedate_variability
        k = np.random.uniform(duedate_tightness - delta_duedate, duedate_tightness + delta_duedate)

        env = simpy.Environment()

        # Start processes and run
        machine_0 = simpy.PriorityResource(env, capacity=1)
        machine_1 = simpy.PriorityResource(env, capacity=1)
        machine_2 = simpy.PriorityResource(env, capacity=1)
        machine_3 = simpy.PriorityResource(env, capacity=1)
        machine_4 = simpy.PriorityResource(env, capacity=1)
        machine_5 = simpy.PriorityResource(env, capacity=1)
        machine_6 = simpy.PriorityResource(env, capacity=1)
        machine_7 = simpy.PriorityResource(env, capacity=1)
        machine_8 = simpy.PriorityResource(env, capacity=1)
        machine_9 = simpy.PriorityResource(env, capacity=1)
        # machine = simpy.Resource(env, capacity=1)

        machine_list = [machine_0, machine_1, machine_2, machine_3, machine_4, machine_5, machine_6, machine_7,
                        machine_8, machine_9]

        for j in range(n_jobs):
            for m in range(number_machines):
                if TASKS[j, m]['prec'] == None:
                    current_machine = machine_list[m]
                    processing_time_new = float(TASKS[j, m]['dur'])
                    job_new = j
                    total_processing_time_current_job = total_processing_time[j]
                    remaining_processing_time_current_job = total_processing_time_current_job
                    release_date_current_job = release_date_list[j]
                    due_date_current_job = due_date_list_jobs[j]
                    env.process(job(env, job_new, current_machine, processing_time=processing_time_new,
                                    total_processing_time=total_processing_time_current_job,
                                    remaining_processing_time=remaining_processing_time_current_job,
                                    due_date=due_date_current_job, release_date=release_date_current_job))

        env.run()

        # for i in range(number_simulations):
        #    env.process(source(env, n_jobs, machine_1))
        #    env.run()

        # Post processing
        # calculate and performance measures of the current simulation run
        total_tardiness = sum(tardiness)
        mean_tardiness = mean(tardiness)
        max_tardiness = max(tardiness)
        number_tardy_jobs = sum(tardy_jobs)
        makespan = max(completion_time)
        # mean_flow_time = mean(flow_time)
        # max_flow_time = max(flow_time)
        # print('Release Time')
        # print(release_time_list)
        # print('processing time')
        # print(processing_time_list)
        # print('Due Dates')
        # print(due_date_list)
        # print(f'Total Tardiness: {total_tardiness}')
        # print(f'Mean Tardiness: {mean_tardiness}')
        # print(f'Max Tardiness: {max_tardiness}')
        # print(f'Number of tardy Jobs: {number_tardy_jobs}')
        # print(completion_time)
        # print(f'Makespan: {makespan}')
        # print(f'Mean flow time: {mean_flow_time}')
        # print(f'Max flow time: {max_flow_time}')
        # print(results)

        # add performance measures of current simulation run to the list for all runs
        avg_makespan_list.append(makespan)
        avg_mean_tardiness_list.append(mean_tardiness)
        avg_max_tardiness_list.append(max_tardiness)
        avg_total_tardiness_list.append(total_tardiness)
        avg_number_tardy_jobs_list.append(number_tardy_jobs)

        #visualize(results)

    # calculate and print the performance measures after all simulation runs
    avg_makespan = mean(avg_makespan_list)
    avg_mean_tardiness = mean(avg_mean_tardiness_list)
    avg_max_tardiness = mean(avg_max_tardiness_list)
    avg_total_tardiness = mean(avg_total_tardiness_list)
    avg_number_tardy_jobs = mean(avg_number_tardy_jobs_list)
    performance = [avg_makespan, avg_number_tardy_jobs, avg_total_tardiness]
    lambda_performance = [0.33, 0.33, 0.33]
    reference_point = reference_point_input
    roh = 0.0001
    fitness = max(lambda_performance[i] * (performance[i] - reference_point[i]) for i in
                  range(0, len(reference_point))) + roh * sum(
        lambda_performance[i] * (performance[i] - reference_point[i]) for i in range(0, len(reference_point)))
    # calculate the fitness value according to the achievement scalarizing function
    if ref_point == None:
        fitness = performance[performance_measure]
    else:
        fitness = max(lambda_performance[i] * (performance[i] - reference_point[i]) for i in
                      range(0, len(reference_point))) + roh * sum(
            lambda_performance[i] * (performance[i] - reference_point[i]) for i in range(0, len(reference_point)))

    # postprocessing
    st.header('Performance measures of the best solution')
    for i in range(len(performance_name)):
        st.subheader(performance_name[i])
        st.subheader(performance[i])
    st.header('Fitness')
    st.write(fitness)
    visualize(results)